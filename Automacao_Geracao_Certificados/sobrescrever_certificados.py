# !/usr/bin/env python3

'''
######################################DESAFIO DESSE SCRIPT#############################################################
Quero ver a possibilidade de criar um programa usando o python para automatizar enviando os dados da planilha para preencher os campos mutáveis no certificado padrão.

Tipo nome do curso , nome participante, tipo de participação, data de inicio, data do final, carga horaria
data da emissão do certificado e as assinaturas do Gestor Geral, do Coordenador
'''

# Canal do Youtube: https://www.youtube.com/@israelalvespro
# Importar as bibliotecas que seram usadas no código!!
import os
import pandas as pd
import openpyxl
from openpyxl import Workbook
from PIL import ImageFont, Image, ImageDraw

# Importar a planilha para começar a trabalhar
tabela = pd.read_excel('planilha_alunos.xlsx')

# ler os dados da planilha
print(tabela)

# Pegar o valor da linha de cada coluna e atribuir a uma variável
table = openpyxl.load_workbook('planilha_alunos.xlsx')
sheet_name = table['Sheet1']

# Para imprimir todos os documentos use essa linha FOR
#for indice, linha in enumerate(sheet_name.iter_rows(min_row=2), start=1):

# Use a linha abaixo para visualizar o funcionamento do código no  começo, imprimindo 5 arquivos
for indice, linha in enumerate(sheet_name.iter_rows(min_row=2, max_row=6), start=1):
    nome_curso = linha[0].value
    nome_participante = linha[1].value
    tipo_participacao = linha[2].value
    data_inicio = linha[3].value
    data_final = linha[4].value
    carga_horaria = str(linha[5].value)
    data_emissao_certificado = linha[6].value

    # Definir as fontes para escrever no documento
    fonte_name = ImageFont.truetype('AnnapurnaSIL-Bold.ttf', 90)
    fonte_geral = ImageFont.truetype('AnnapurnaSIL-Regular.ttf', 80)
    fonte_data = ImageFont.truetype('AnnapurnaSIL-Regular.ttf', 55)

    # Abrir a imagem para poder sobrescrever
    try:
        imagem = Image.open('certificado_padrao.jpg')
        desenhar = ImageDraw.Draw(imagem)
    except FileNotFoundError:
        print(f"Arquivo de imagem do certificado não encontrado")

    # Definir parâmetros para poder desenhar nos espaços corretos da imagem
    desenhar.text((1070, 957), nome_curso, fill='black', font=fonte_geral)
    desenhar.text((1020, 827), nome_participante, fill='black', font=fonte_name)
    desenhar.text((1436, 1070), tipo_participacao, fill='black', font=fonte_geral)
    desenhar.text((750, 1777), data_inicio, fill='black', font=fonte_data)
    desenhar.text((750, 1940), data_final, fill='black', font=fonte_data)
    desenhar.text((1496, 1180), carga_horaria, fill='black', font=fonte_name)
    desenhar.text((2220, 1940), data_emissao_certificado, fill='black', font=fonte_data)


    # Salvando a imagem alterada
    def salvar_arquivo(nome_arquivo):
        try:
            if os.path.exists(nome_arquivo):
                print(f"Arquivo: {nome_arquivo} já existe")

            else:
                # Salvando a imagem alterada
                imagem.save(nome_arquivo)
                print(f"Imagem salva como {nome_arquivo}")

        except OSError as e:
            print(f"Erro ao salvar a imagem: {e}")


    # Definindo o valor para nome do arquivo e chamando a função salvar_arquivo
    nome_arquivo = f"({indice}) {nome_participante} certificado.png"
    salvar_arquivo(nome_arquivo)
